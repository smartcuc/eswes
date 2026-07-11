#####################
# energy/api/views.py
#####################

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from zoneinfo import ZoneInfo
from django.utils import timezone
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from devices.models import Device

from energy.services.energy import get_energy_data
from energy.services.charts import (get_chart_data,)
from energy.ems.models import EMSSignalSource

from openpyxl import Workbook
from openpyxl.styles import Font

import csv

from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)
from reportlab.lib.styles import getSampleStyleSheet
import matplotlib.pyplot as plt


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dashboard_me(request):
    data = get_energy_data(request.user)
    return Response(data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def chart_data(request):

    metric = request.GET.get("metric")
    period = request.GET.get(
        "period",
        "24h",
    )

    if period not in [
        "1h",
        "6h",
        "24h",
        "5d",
    ]:
        return Response(
            {"detail": "invalid period"},
            status=400,
        )

    if metric not in [
        "load",
        "pv",
        "grid",
        "today",
    ]:
        return Response(
            {"detail": "invalid metric"},
            status=400,
        )

    if metric == "pv":

        device_ids = list(
            EMSSignalSource.objects.filter(
                home__user=request.user,
                signal_type="pv",
            ).values_list(
                "device_id",
                flat=True,
            )
        )

    else:

        device_ids = list(
            EMSSignalSource.objects.filter(
                home__user=request.user,
                signal_type="grid",
            ).values_list(
                "device_id",
                flat=True,
            )
        )

    home = request.user.homes.first()
    timezone_name = home.timezone if home else "UTC"

    data = get_chart_data(
        device_ids,
        period,
        timezone_name=timezone_name,
    )

    return Response(data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def configure_device(request, device_id):
    device = get_object_or_404(Device, id=device_id, user=request.user)

    device.role = request.data.get("role")
    device.room = request.data.get("room")
    device.name = request.data.get("name", device.name)

    device.configured = True
    device.save()

    return Response({"status": "ok"})

# ----------------
# Export - Excel
# ----------------

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_chart_xlsx(request):

    metric = request.GET.get("metric")
    period = request.GET.get("period", "24h")

    if metric == "pv":

        device_ids = list(
            EMSSignalSource.objects.filter(
                home__user=request.user,
                signal_type="pv",
            ).values_list(
                "device_id",
                flat=True,
            )
        )

    else:

        device_ids = list(
            EMSSignalSource.objects.filter(
                home__user=request.user,
                signal_type="grid",
            ).values_list(
                "device_id",
                flat=True,
            )
        )

    home = request.user.homes.first()
    timezone_name = home.timezone if home else "UTC"

    export_time = (
        timezone.now().astimezone(ZoneInfo(timezone_name)).strftime("%d.%m.%Y %H:%M")
    )

    data = get_chart_data(
        device_ids,
        period,
        timezone_name=timezone_name,
    )

    wb = Workbook()
    ws = wb.active

    ws.title = f"{metric}_{period}"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18

    ws["A1"].font = Font(
        bold=True,
        size=16,
    )

    metric_labels = {
        "load": "Bedarf",
        "pv": "Erzeugung",
        "grid": "Bezug/Einspeisung",
        "today": "Tagesverbrauch",
    }

    ws["A8"].font = Font(bold=True)
    ws["B8"].font = Font(bold=True)

    ws["A1"] = "Sharegy Energieexport"

    ws["A3"] = "Metrik"
    ws["B3"] = metric_labels.get(
        metric,
        metric,
    )

    ws["A4"] = "Zeitraum"
    ws["B4"] = period

    ws["A5"] = "Einheit"
    ws["B5"] = data["unit"]

    ws["A6"] = "Exportiert"
    ws["B6"] = export_time

    # Leerzeile
    ws.append([])

    ws["A8"] = "Zeitpunkt"
    ws["B8"] = f"Wert ({data['unit']})"

    row = 9

    for ts, value in zip(
        data["export_timestamps"],
        data["values"],
    ):
        ws.cell(row=row, column=1, value=ts)
        ws.cell(row=row, column=2, value=value)
        row += 1

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )

    metric_label = metric_labels.get(
        metric,
        metric,
    )

    safe_label = metric_label.replace("/", "-").replace("\\", "-").replace(" ", "_")

    response["Content-Disposition"] = f'attachment; filename="sharegy_{safe_label}_{period}.xlsx"'

    wb.save(response)

    return response


# ----------------
# Export - CSV
# ----------------

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_chart_csv(request):

    metric = request.GET.get("metric")
    period = request.GET.get("period", "24h")

    if metric == "pv":

        device_ids = list(
            EMSSignalSource.objects.filter(
                home__user=request.user,
                signal_type="pv",
            ).values_list(
                "device_id",
                flat=True,
            )
        )

    else:

        device_ids = list(
            EMSSignalSource.objects.filter(
                home__user=request.user,
                signal_type="grid",
            ).values_list(
                "device_id",
                flat=True,
            )
        )

    home = request.user.homes.first()
    timezone_name = home.timezone if home else "UTC"

    export_time = (
        timezone.now().astimezone(ZoneInfo(timezone_name)).strftime("%d.%m.%Y %H:%M")
    )

    data = get_chart_data(
        device_ids,
        period,
        timezone_name=timezone_name,
    )

    metric_labels = {
        "load": "Bedarf",
        "pv": "Erzeugung",
        "grid": "Bezug/Einspeisung",
        "today": "Tagesverbrauch",
    }

    response = HttpResponse(
        content_type="text/csv"
    )

    metric_label = metric_labels.get(
        metric,
        metric,
    )

    safe_label = metric_label.replace("/", "-").replace("\\", "-").replace(" ", "_")

    response["Content-Disposition"] = (
        f'attachment; filename="sharegy_{safe_label}_{period}.csv"'
    )

    writer = csv.writer(
        response,
        delimiter=";"
    )

    writer.writerow(["Sharegy Energieexport"])
    writer.writerow([])

    writer.writerow(["Metrik", metric_labels.get(metric, metric)])
    writer.writerow(["Zeitraum", period])
    writer.writerow(["Einheit", data["unit"]])
    writer.writerow(["Exportiert", export_time])

    writer.writerow([])

    writer.writerow([
        "Zeitpunkt",
        f"Wert ({data['unit']})"
    ])

    for ts, value in zip(
        data["export_timestamps"],
        data["values"],
    ):
        writer.writerow(
            [
                ts,
                str(value).replace(".", ","),
            ]
        )

    return response


# ----------------
# Export - PDF
# ----------------
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_chart_pdf(request):

    metric = request.GET.get("metric")
    period = request.GET.get("period", "24h")

    if metric == "pv":

        device_ids = list(
            EMSSignalSource.objects.filter(
                home__user=request.user,
                signal_type="pv",
            ).values_list(
                "device_id",
                flat=True,
            )
        )

    else:

        device_ids = list(
            EMSSignalSource.objects.filter(
                home__user=request.user,
                signal_type="grid",
            ).values_list(
                "device_id",
                flat=True,
            )
        )

    home = request.user.homes.first()
    timezone_name = home.timezone if home else "UTC"

    export_time = (
        timezone.now().astimezone(ZoneInfo(timezone_name)).strftime("%d.%m.%Y %H:%M")
    )

    data = get_chart_data(
        device_ids,
        period,
        timezone_name=timezone_name,
    )

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
    )

    styles = getSampleStyleSheet()

    elements = []

    title = Paragraph(
        "<font size='22'><b>Sharegy</b></font><br/>"
        "<font size='16'>Energieexport</font>",
        styles["Title"],
    )

    elements.append(title)

    elements.append(Spacer(1, 20))

    metric_labels = {
        "load": "Bedarf",
        "pv": "Erzeugung",
        "grid": "Bezug/Einspeisung",
        "today": "Tagesverbrauch",
    }

    info_table = Table(
        [
            ["Metrik", metric_labels.get(metric, metric)],
            ["Zeitraum", period],
            ["Einheit", data["unit"]],
            ["Zeitzone", timezone_name],
            ["Exportiert", export_time],
        ],
        colWidths=[100, 250],
    )

    info_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#D1D5DB")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    elements.append(info_table)
    elements.append(Spacer(1, 30))

    elements.append(
        Paragraph(
            "Leistungsverlauf",
            styles["Heading2"],
        )
    )

    elements.append(Spacer(1, 8))

    chart_buffer = BytesIO()

    timestamps = [
        ts[11:16]
        for ts in data["export_timestamps"]
    ]

    plt.figure(figsize=(8, 3))
    plt.margins(x=0)

    plt.plot(
        data["values"],
        linewidth=2.5,
        color="#2563EB",
    )

    plt.fill_between(
        range(len(data["values"])),
        data["values"],
        alpha=0.15,
        color="#2563EB",
    )

    plt.title(metric_labels.get(metric, metric))

    plt.suptitle(
        f"Zeitraum: {period}",
        fontsize=9,
    )

    step = max(1, len(timestamps) // 5)

    plt.xticks(
        range(0, len(timestamps), step),
        timestamps[::step],
        rotation=45,
    )

    plt.ylabel(data["unit"])
    plt.xlabel("Zeit")

    plt.grid(
        True,
        alpha=0.3,
    )

    plt.tight_layout()

    plt.savefig(
        chart_buffer,
        format="png",
    )

    plt.close()

    chart_buffer.seek(0)

    chart = Image(
        chart_buffer,
        width=450,
        height=220,
    )

    elements.append(chart)
    elements.append(Spacer(1, 25))

    elements.append(
        Paragraph(
            "Zusammenfassung",
            styles["Heading2"],
        )
    )

    elements.append(Spacer(1, 8))

    values = data["values"]

    max_value = max(values) if values else 0
    min_value = min(values) if values else 0
    avg_value = sum(values) / len(values) if values else 0
    count_value = len(values)

    summary_table = Table(
        [
            ["Maximum", f"{max_value:.1f} {data['unit']}"],
            ["Minimum", f"{min_value:.1f} {data['unit']}"],
            ["Durchschnitt", f"{avg_value:.1f} {data['unit']}"],
            ["Messwerte", str(count_value)],
        ],
        colWidths=[120, 180],
    )

    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#D1D5DB")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    elements.append(summary_table)
    elements.append(Spacer(1, 25))

    table_data = [["Zeitpunkt", f"Wert ({data['unit']})"]]

    for ts, value in zip(
        data["export_timestamps"],
        data["values"],
    ):
        table_data.append([
            ts,
            str(value).replace(".", ","),
        ])

    data_table = Table(
        table_data,
        colWidths=[220, 100],
    )

    data_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ]
        )
    )

    for i in range(1, len(table_data)):
        if i % 2 == 0:
            data_table.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, i),
                            (-1, i),
                            colors.HexColor("#F9FAFB"),
                        )
                    ]
                )
            )

    elements.append(data_table)

    doc.build(elements)

    response = HttpResponse(content_type="application/pdf")

    metric_label = metric_labels.get(
        metric,
        metric,
    )

    safe_label = metric_label.replace("/", "-").replace("\\", "-").replace(" ", "_")

    response["Content-Disposition"] = (
        f'attachment; filename="sharegy_{safe_label}_{period}.pdf"'
    )

    buffer.seek(0)
    response.write(buffer.getvalue())

    return response
